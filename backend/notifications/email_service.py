"""
Weekly GW email report via SendGrid.

Sends a styled "coach's briefing" HTML email every Friday 10am (or on demand).
Attaches an Excel decision journal (openpyxl) with squad + GW stats.

Only active if SENDGRID_API_KEY and NOTIFICATION_TO_EMAIL are set.
"""
import io
from datetime import datetime
from loguru import logger

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class EmailService:
    def __init__(self):
        from core.config import settings
        self.settings = settings

        if not settings.email_enabled:
            raise RuntimeError("Email not configured (missing SENDGRID_API_KEY or NOTIFICATION_TO_EMAIL)")

        import sendgrid
        self.sg = sendgrid.SendGridAPIClient(api_key=settings.SENDGRID_API_KEY)

    async def send_weekly_report(self, gw_data: dict | None = None) -> bool:
        """
        Send weekly GW brief email with Excel attachment.
        gw_data: pre-built dict from /api/intel/gw — if None, skip attachment.
        """
        import asyncio
        from functools import partial

        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(
            None, partial(self._send_sync, gw_data)
        )
        return result

    def _send_sync(self, gw_data: dict | None) -> bool:
        """Sync SendGrid send (run in executor)."""
        try:
            from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
            import base64

            gw_num = gw_data.get("gameweek", "?") if gw_data else "?"
            subject = f"GW{gw_num} Strategy Brief — FPL Intelligence"

            html_body = self._build_html(gw_data)

            message = Mail(
                from_email=self.settings.SENDGRID_FROM_EMAIL or "fpl@intelligence.local",
                to_emails=self.settings.NOTIFICATION_TO_EMAIL,
                subject=subject,
                html_content=html_body,
            )

            # Excel attachment
            if gw_data:
                excel_bytes = self._build_excel(gw_data)
                encoded = base64.b64encode(excel_bytes).decode()
                attachment = Attachment(
                    file_content=FileContent(encoded),
                    file_name=FileName(f"fpl_gw{gw_num}_brief.xlsx"),
                    file_type=FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    disposition=Disposition("attachment"),
                )
                message.attachment = attachment

            response = self.sg.send(message)
            logger.info(f"Weekly email sent (GW{gw_num}), status={response.status_code}")
            return response.status_code in (200, 202)

        except Exception as e:
            logger.error(f"Email send failed: {e}")
            return False

    def _build_html(self, gw_data: dict | None) -> str:
        """Build a clean HTML coach's briefing email."""
        if not gw_data:
            return "<p>FPL Intelligence Engine — weekly report unavailable.</p>"

        gw = gw_data.get("gameweek", "?")
        captain = gw_data.get("captain_recommendation", {})
        injuries = gw_data.get("injury_alerts", [])
        suspensions = gw_data.get("suspension_risk", [])
        blanks = gw_data.get("blank_gw_starters", [])
        doubles = gw_data.get("double_gw_players", [])

        captain_html = ""
        if captain:
            captain_html = f"""
            <tr>
              <td style="padding:8px 0;color:#6b7280;font-size:14px;">Captain Pick</td>
              <td style="padding:8px 0;font-weight:600;font-size:14px;">
                {captain.get('web_name','?')} — Score: {captain.get('score', 0):.1f}
                (FDR {captain.get('fdr_next','?')}, {'Home' if captain.get('is_home_next') else 'Away'})
              </td>
            </tr>"""

        injury_rows = "".join(
            f"<li>{i['web_name']} — {i.get('news','')[:80]}</li>"
            for i in injuries[:5]
        )

        suspension_rows = "".join(
            f"<li>⚠️ {s['web_name']} ({s.get('yellow_cards',0)} yellows — SELL BEFORE CUTOFF)</li>"
            for s in suspensions
        )

        dgw_rows = "".join(
            f"<li>{d['web_name']} — {d.get('predicted_xpts_next',0):.1f} xPts (DGW)</li>"
            for d in doubles[:5]
        )

        return f"""
<!DOCTYPE html>
<html>
<head>
  <style>
    body {{ font-family: 'Georgia', serif; background:#f9f7f4; margin:0; padding:0; }}
    .container {{ max-width:600px; margin:0 auto; background:#fff; border:1px solid #e5e7eb; }}
    .header {{ background:#1a3a2a; color:#fff; padding:24px 32px; }}
    .header h1 {{ margin:0; font-size:22px; letter-spacing:1px; }}
    .header p {{ margin:4px 0 0; color:#86efac; font-size:14px; }}
    .body {{ padding:32px; }}
    .section {{ margin-bottom:24px; }}
    .section h2 {{ font-size:16px; color:#1a3a2a; border-bottom:2px solid #86efac; padding-bottom:6px; margin-bottom:12px; }}
    table {{ width:100%; border-collapse:collapse; }}
    .alert {{ background:#fef3c7; border-left:4px solid #f59e0b; padding:10px 14px; margin:8px 0; border-radius:2px; font-size:13px; }}
    .danger {{ background:#fee2e2; border-left:4px solid #ef4444; }}
    .good {{ background:#d1fae5; border-left:4px solid #10b981; }}
    .footer {{ background:#f3f4f6; padding:16px 32px; font-size:12px; color:#6b7280; }}
  </style>
</head>
<body>
<div class="container">
  <div class="header">
    <h1>⚽ GW{gw} Strategy Brief</h1>
    <p>FPL Intelligence Engine — Coach's Briefing</p>
  </div>
  <div class="body">

    <div class="section">
      <h2>This Week's Decisions</h2>
      <table>
        {captain_html}
      </table>
    </div>

    {"<div class='section'><h2>Double Gameweek Targets</h2><ul>" + dgw_rows + "</ul></div>" if dgw_rows else ""}

    {"<div class='section'><h2>🏥 Injury Alerts (Your Squad)</h2><ul>" + injury_rows + "</ul></div>" if injury_rows else ""}

    {"<div class='section'><h2>🟨 Suspension Risk</h2><ul class='danger'>" + suspension_rows + "</ul></div>" if suspension_rows else ""}

    {"<div class='section'><h2>❌ Blank GW Starters</h2><div class='alert'>" + ", ".join(b['web_name'] for b in blanks) + " have no fixture this GW. Consider selling or benching.</div></div>" if blanks else ""}

    <div class="section">
      <h2>Quick Reminder</h2>
      <div class="alert">Log in to FPL and confirm your squad before the deadline. Check the app for live optimisation suggestions.</div>
    </div>
  </div>
  <div class="footer">
    Generated by FPL Intelligence Engine · {datetime.now().strftime('%d %b %Y %H:%M')} ·
    Not financial advice — use your judgement!
  </div>
</div>
</body>
</html>"""

    # ── Pre-deadline alert ─────────────────────────────────────────────────────

    async def send_deadline_alert(
        self,
        to_email: str,
        gw_id: int,
        intel_data: dict | None = None,
    ) -> bool:
        """
        Send a rich HTML pre-deadline alert 24h before GW kickoff.
        intel_data: response from GET /api/intel/gw (captain, transfers, injuries, chips).
        """
        import asyncio
        from functools import partial

        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(
            None, partial(self._send_deadline_sync, to_email, gw_id, intel_data)
        )

    def _send_deadline_sync(
        self,
        to_email: str,
        gw_id: int,
        intel_data: dict | None,
    ) -> bool:
        try:
            from sendgrid.helpers.mail import Mail
            html_body = self._build_deadline_html(gw_id, intel_data or {})
            message = Mail(
                from_email=self.settings.SENDGRID_FROM_EMAIL or "fpl@intelligence.local",
                to_emails=to_email,
                subject=f"⏰ GW{gw_id} Deadline Tomorrow — Your FPL Briefing",
                html_content=html_body,
            )
            response = self.sg.send(message)
            ok = response.status_code in (200, 202)
            logger.info(f"Deadline alert GW{gw_id} → {to_email}: status={response.status_code}")
            return ok
        except Exception as e:
            logger.error(f"Deadline alert send failed (GW{gw_id} → {to_email}): {e}")
            return False

    def _build_deadline_html(self, gw_id: int, intel: dict) -> str:
        """
        Rich HTML pre-deadline briefing email.
        Sections: header, captain card, top transfers, injury alerts, chip card, footer.
        All CSS is inline for maximum email client compatibility.
        """
        captain = intel.get("captain_recommendation") or {}
        transfers = intel.get("suggested_transfers") or []
        injuries = intel.get("injury_alerts") or []
        chip = intel.get("chip_recommendation") or {}
        deadline_str = intel.get("deadline_time", "")
        if deadline_str:
            try:
                from datetime import timezone
                dt = datetime.fromisoformat(deadline_str.replace("Z", "+00:00"))
                deadline_fmt = dt.strftime("%-d %b %Y, %H:%M UTC")
            except Exception:
                deadline_fmt = deadline_str
        else:
            deadline_fmt = "Soon"

        # ── Captain card ──────────────────────────────────────────────────────
        cap_html = ""
        if captain:
            cap_name = captain.get("web_name", "—")
            cap_xpts = captain.get("score") or captain.get("xpts") or 0
            cap_reason = captain.get("reasoning") or captain.get("rationale") or ""
            fixture = captain.get("fixture", "")
            cap_html = f"""
        <div style="background:#0D2B1A;border-radius:10px;padding:20px 24px;margin-bottom:20px;">
          <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px;">
            <div>
              <p style="margin:0;color:#86efac;font-size:12px;letter-spacing:1px;text-transform:uppercase;font-family:sans-serif;">Captain Pick</p>
              <p style="margin:6px 0 0;color:#ffffff;font-size:26px;font-weight:700;font-family:Georgia,serif;">{cap_name}</p>
              {f'<p style="margin:4px 0 0;color:#6ee7b7;font-size:13px;font-family:sans-serif;">{fixture}</p>' if fixture else ''}
            </div>
            <div style="background:#16a34a;border-radius:24px;padding:6px 18px;text-align:center;">
              <p style="margin:0;color:#fff;font-size:22px;font-weight:700;font-family:sans-serif;">{float(cap_xpts):.1f}</p>
              <p style="margin:0;color:#bbf7d0;font-size:11px;font-family:sans-serif;">xPts (2×)</p>
            </div>
          </div>
          {f'<p style="margin:12px 0 0;color:#d1fae5;font-size:13px;font-family:sans-serif;line-height:1.5;">💡 {cap_reason[:180]}</p>' if cap_reason else ''}
        </div>"""

        # ── Transfers card ────────────────────────────────────────────────────
        transfer_rows = ""
        for t in transfers[:3]:
            out_name = t.get("out", {}).get("web_name") or t.get("player_out", "?")
            in_name = t.get("in", {}).get("web_name") or t.get("player_in", "?")
            delta = t.get("xpts_gain") or t.get("delta") or 0
            confidence = t.get("confidence") or t.get("score") or 0
            transfer_rows += f"""
          <tr>
            <td style="padding:10px 8px;border-bottom:1px solid #e5e7eb;">
              <span style="background:#fee2e2;color:#b91c1c;border-radius:4px;padding:2px 8px;font-size:12px;font-weight:600;font-family:sans-serif;">OUT</span>
              <span style="margin-left:8px;font-size:14px;color:#111827;font-family:sans-serif;">{out_name}</span>
            </td>
            <td style="padding:10px 8px;border-bottom:1px solid #e5e7eb;text-align:center;color:#6b7280;font-size:18px;">→</td>
            <td style="padding:10px 8px;border-bottom:1px solid #e5e7eb;">
              <span style="background:#d1fae5;color:#065f46;border-radius:4px;padding:2px 8px;font-size:12px;font-weight:600;font-family:sans-serif;">IN</span>
              <span style="margin-left:8px;font-size:14px;color:#111827;font-family:sans-serif;">{in_name}</span>
            </td>
            <td style="padding:10px 8px;border-bottom:1px solid #e5e7eb;text-align:right;font-size:13px;font-family:sans-serif;">
              <span style="color:#16a34a;font-weight:600;">+{float(delta):.1f} xPts</span>
              {f'<br><span style="color:#6b7280;">{float(confidence):.0f}% confidence</span>' if confidence else ''}
            </td>
          </tr>"""

        transfers_html = ""
        if transfer_rows:
            transfers_html = f"""
        <div style="margin-bottom:20px;">
          <h2 style="font-size:14px;color:#374151;text-transform:uppercase;letter-spacing:1px;margin:0 0 12px;font-family:sans-serif;border-bottom:2px solid #d1fae5;padding-bottom:6px;">🔄 Suggested Transfers</h2>
          <table style="width:100%;border-collapse:collapse;">{transfer_rows}</table>
        </div>"""

        # ── Injury alerts ─────────────────────────────────────────────────────
        injury_items = ""
        for inj in injuries[:5]:
            name = inj.get("web_name", "?")
            news = inj.get("news", "")[:100]
            injury_items += f"""
          <div style="display:flex;align-items:flex-start;padding:8px 0;border-bottom:1px solid #fee2e2;">
            <span style="color:#ef4444;font-size:16px;margin-right:10px;flex-shrink:0;">●</span>
            <div>
              <span style="font-weight:600;color:#111827;font-size:14px;font-family:sans-serif;">{name}</span>
              {f'<br><span style="color:#6b7280;font-size:12px;font-family:sans-serif;">{news}</span>' if news else ''}
            </div>
          </div>"""

        injuries_html = ""
        if injury_items:
            injuries_html = f"""
        <div style="background:#fff5f5;border:1px solid #fecaca;border-radius:8px;padding:16px;margin-bottom:20px;">
          <h2 style="font-size:14px;color:#991b1b;text-transform:uppercase;letter-spacing:1px;margin:0 0 10px;font-family:sans-serif;">🏥 Injury Alerts</h2>
          {injury_items}
        </div>"""

        # ── Chip recommendation ───────────────────────────────────────────────
        chip_html = ""
        if chip and chip.get("recommended"):
            chip_name = chip.get("chip_name") or chip.get("recommended") or "Chip"
            chip_reason = chip.get("reasoning") or chip.get("rationale") or ""
            chip_html = f"""
        <div style="background:#fffbeb;border:1px solid #fbbf24;border-radius:8px;padding:16px;margin-bottom:20px;">
          <h2 style="font-size:14px;color:#92400e;text-transform:uppercase;letter-spacing:1px;margin:0 0 8px;font-family:sans-serif;">⚡ Chip Recommendation</h2>
          <p style="margin:0;font-weight:600;color:#78350f;font-size:15px;font-family:sans-serif;">{chip_name}</p>
          {f'<p style="margin:6px 0 0;color:#92400e;font-size:13px;font-family:sans-serif;">{chip_reason[:200]}</p>' if chip_reason else ''}
        </div>"""

        unsubscribe_url = f"http://localhost:8000/api/user/profile?team_id=0"  # placeholder

        return f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f3f4f6;font-family:Georgia,serif;">
<div style="max-width:600px;margin:0 auto;background:#ffffff;border-radius:12px;overflow:hidden;box-shadow:0 4px 16px rgba(0,0,0,0.08);">

  <!-- Header -->
  <div style="background:#0D2B1A;padding:28px 32px;">
    <p style="margin:0;color:#86efac;font-size:11px;letter-spacing:2px;text-transform:uppercase;font-family:sans-serif;">FPL Intelligence Engine</p>
    <h1 style="margin:8px 0 0;color:#ffffff;font-size:24px;font-weight:700;">GW{gw_id} Deadline in 24h ⏰</h1>
    <p style="margin:6px 0 0;color:#6ee7b7;font-size:13px;font-family:sans-serif;">Deadline: {deadline_fmt}</p>
  </div>

  <!-- Body -->
  <div style="padding:28px 32px;">
    {cap_html}
    {transfers_html}
    {injuries_html}
    {chip_html}

    <div style="background:#f0fdf4;border-radius:8px;padding:14px 18px;text-align:center;">
      <p style="margin:0;color:#166534;font-size:13px;font-family:sans-serif;">
        ⚽ Make your changes at <a href="https://fantasy.premierleague.com" style="color:#16a34a;font-weight:600;">fantasy.premierleague.com</a> before the deadline.
      </p>
    </div>
  </div>

  <!-- Footer -->
  <div style="background:#f9fafb;border-top:1px solid #e5e7eb;padding:16px 32px;">
    <p style="margin:0;font-size:11px;color:#9ca3af;font-family:sans-serif;text-align:center;">
      Generated by FPL Intelligence Engine · {datetime.now().strftime('%d %b %Y %H:%M')} ·
      <a href="{unsubscribe_url}" style="color:#9ca3af;">Manage alerts</a>
    </p>
  </div>

</div>
</body>
</html>"""

    # ── Admin pipeline-failure alert ──────────────────────────────────────────

    async def send_admin_alert(self, subject: str, body: str) -> bool:
        """
        Send a plain-text alert to the configured ADMIN_ALERT_EMAIL address.

        Used by the scheduler to notify the operator when a critical job fails.
        Does nothing (returns False) when ADMIN_ALERT_EMAIL is not configured or
        when the email service itself is not set up.
        """
        admin_email = self.settings.ADMIN_ALERT_EMAIL
        if not admin_email:
            return False

        import asyncio
        from functools import partial

        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(
            None, partial(self._send_admin_sync, subject, body, admin_email)
        )

    def _send_admin_sync(self, subject: str, body: str, admin_email: str) -> bool:
        try:
            from sendgrid.helpers.mail import Mail

            ts = datetime.now().strftime("%d %b %Y %H:%M UTC")
            html_body = f"""<!DOCTYPE html>
<html>
<head><style>
body{{font-family:monospace;background:#0d1117;color:#e6edf3;padding:24px;}}
.box{{background:#161b22;border:1px solid #30363d;border-radius:6px;padding:16px 20px;max-width:600px;}}
h2{{color:#f85149;margin:0 0 12px;font-size:16px;}}
pre{{color:#c9d1d9;font-size:13px;white-space:pre-wrap;word-break:break-word;background:#0d1117;
    padding:12px;border-radius:4px;border:1px solid #30363d;margin:0;}}
small{{color:#6e7681;font-size:11px;}}
</style></head>
<body><div class="box">
<h2>🚨 FPL Intelligence — Pipeline Alert</h2>
<pre>{body}</pre>
<br><small>Sent: {ts} · FPL Intelligence Engine</small>
</div></body>
</html>"""
            message = Mail(
                from_email=self.settings.SENDGRID_FROM_EMAIL or "fpl@intelligence.local",
                to_emails=admin_email,
                subject=f"[FPL] {subject}",
                html_content=html_body,
            )
            response = self.sg.send(message)
            ok = response.status_code in (200, 202)
            logger.info(f"Admin alert sent → {admin_email}: status={response.status_code}")
            return ok
        except Exception as e:
            logger.error(f"Admin alert send failed (to={admin_email}): {e}")
            return False

    def _build_excel(self, gw_data: dict) -> bytes:
        """Build an Excel decision journal for the current GW."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"GW{gw_data.get('gameweek','?')} Brief"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A3A2A", end_color="1A3A2A", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            "Category", "Player / Detail", "Metric", "Value", "Action / Note"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        row = 2

        def add_row(category: str, player: str, metric: str, value, note: str = ""):
            nonlocal row
            data = [category, player, metric, value, note]
            fills = {
                "Captain": "D1FAE5",
                "Injury": "FEF3C7",
                "Suspension": "FEE2E2",
                "Double GW": "EDE9FE",
            }
            fill_color = fills.get(category, "FFFFFF")
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            row += 1

        # Captain
        cap = gw_data.get("captain_recommendation")
        if cap:
            add_row("Captain", cap.get("web_name", ""), "xPts Score", round(cap.get("score", 0), 2), "Recommended captain")

        # Injuries
        for inj in gw_data.get("injury_alerts", []):
            add_row("Injury", inj["web_name"], "Status", inj.get("status", "?"), inj.get("news", "")[:100])

        # Suspensions
        for sus in gw_data.get("suspension_risk", []):
            add_row("Suspension", sus["web_name"], "Yellow Cards", sus.get("yellow_cards", 0), "SELL before cutoff")

        # Double GW
        for dgw in gw_data.get("double_gw_players", []):
            add_row("Double GW", dgw["web_name"], "xPts (DGW)", round(dgw.get("predicted_xpts_next", 0), 1), "Two fixtures this GW")

        # Blank GW
        for blk in gw_data.get("blank_gw_starters", []):
            add_row("Blank GW", blk["web_name"], "Fixtures", 0, "No fixture — bench or sell")

        # Column widths
        col_widths = [15, 22, 18, 12, 40]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 20

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
